#sync_logs_to_excel.py
import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def sync_json_to_excel(
    json_path="src/live_signals_log.json", excel_path="paper_trading_tracker.xlsx"
):
    # 1. Read and parse the live JSON log
    if not os.path.exists(json_path):
        print(f"  [Sync] ERROR: Could not find '{json_path}'")
        return

    with open(json_path, "r") as f:
        try:
            raw_data = json.load(f)
        except json.JSONDecodeError:
            print("  [Sync] ERROR: JSON file is empty or corrupted.")
            return

    # Filter out scan summaries, keeping only valid trade signals
    trade_signals = [item for item in raw_data if "type" not in item]
    if not trade_signals:
        print("  [Sync] No active trade signals found in the JSON log yet.")
        return

    # Convert incoming JSON signals into a clean DataFrame
    incoming_rows = []
    for sig in trade_signals:
        incoming_rows.append(
            {
                "scan_time": sig.get("scan_time"),
                "symbol": sig.get("symbol"),
                "direction": sig.get("direction"),
                "confidence": sig.get("confidence"),
                "confidence_band": sig.get("confidence_band"),
                "entry_price": sig.get("entry_price"),
                "stop_loss": sig.get("stop_loss"),
                "take_profit": sig.get("take_profit"),
                "actual_fill": "",  # Manual Field placeholder
                "outcome": "OPEN",  # Default state
                "actual_exit": "",  # Manual Field placeholder
                "r_multiple": "",  # Formula controlled
                "slippage_r": "",  # Formula controlled
                "notes": "",  # Manual Field placeholder
            }
        )
    df_new = pd.DataFrame(incoming_rows)

    # 2. Look for existing data to prevent overwriting your manual work
    # Columns that belong to the user - never overwrite with data from JSON
    MANUAL_COLUMNS = ["actual_fill", "outcome", "actual_exit", "r_multiple", "slippage_r", "notes"]

    if os.path.exists(excel_path):
        try:
            df_existing = pd.read_excel(excel_path, sheet_name="Trades", dtype=str)
            df_existing = df_existing.fillna("")

            df_new.set_index(["scan_time", "symbol"], inplace=True)
            df_existing.set_index(["scan_time", "symbol"], inplace=True)

            # Start from df_new (has all current signals with correct system fields)
            df_final = df_new.copy()

            # For rows that already exist in df_existing:
            # - Keep all system fields from df_new (entry_price, stop_loss etc may update)
            # - Restore manual fields from df_existing (never overwrite user input)
            common_idx = df_final.index.intersection(df_existing.index)
            for col in MANUAL_COLUMNS:
                if col in df_existing.columns and col in df_final.columns:
                    df_final.loc[common_idx, col] = df_existing.loc[common_idx, col]

            # Append rows that exist in df_existing but NOT in df_new
            # (older signals no longer in JSON - preserve them)
            old_only_idx = df_existing.index.difference(df_final.index)
            if not old_only_idx.empty:
                df_final = pd.concat([df_final, df_existing.loc[old_only_idx]])

            df_final = df_final.reset_index()

        except Exception as e:
            print(f"  [Sync] Warning: could not merge existing file ({e}). Creating fresh sheet.")
            df_final = df_new.reset_index() if df_new.index.names[0] else df_new
    else:
        # First run - set outcome default to OPEN for all new signals
        df_new["outcome"] = "OPEN"
        df_final = df_new
        
    df_final["outcome"] = df_final["outcome"].replace("", "OPEN").fillna("OPEN")

    # Guarantee strict column order matching the spreadsheet design
    columns_order = [
        "scan_time",
        "symbol",
        "direction",
        "confidence",
        "confidence_band",
        "entry_price",
        "stop_loss",
        "take_profit",
        "actual_fill",
        "outcome",
        "actual_exit",
        "r_multiple",
        "slippage_r",
        "notes",
    ]
    df_final = df_final[columns_order]

    # 3. Rebuild the Excel Workbook with exact Copilot formulas
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Write the data framework first
        df_final.to_excel(writer, sheet_name="Trades", index=False)

        # Build Summary Data Structure
        summary_data = [
            ["Metric", "Value"],
            [
                "Total signals",
                '=COUNTA(Trades!A2:A1000)-COUNTIF(Trades!A2:A1000, "")',
            ],
            [
                "Closed trades",
                '=COUNTIF(Trades!J2:J1000,"WIN")+COUNTIF(Trades!J2:J1000,"LOSS")',
            ],
            ["Win rate", '=IFERROR(COUNTIF(Trades!J2:J1000,"WIN")/B3,"")'],
            [
                "Avg R",
                '=IFERROR(AVERAGEIF(Trades!J2:J1000,"<>OPEN",Trades!L2:L1000),"")',
            ],
            [
                "Total R",
                '=IFERROR(SUMIF(Trades!J2:J1000,"<>OPEN",Trades!L2:L1000),"")',
            ],
            ["Avg slippage R", '=IFERROR(AVERAGE(Trades!M2:M1000),"")'],
            ["By band", ""],
            ["70-74 count", '=COUNTIF(Trades!E2:E1000,"70-74")'],
            [
                "70-74 WR",
                '=IFERROR(COUNTIFS(Trades!E2:E1000,"70-74",Trades!J2:J1000,"WIN")/COUNTIFS(Trades!E2:E1000,"70-74",Trades!J2:J1000,"<>OPEN"),"")',
            ],
            [
                "70-74 avg R",
                '=IFERROR(AVERAGEIFS(Trades!L2:L1000,Trades!E2:E1000,"70-74",Trades!J2:J1000,"<>OPEN"),"")',
            ],
            ["80-84 count", '=COUNTIF(Trades!E2:E1000,"80-84")'],
            [
                "80-84 WR",
                '=IFERROR(COUNTIFS(Trades!E2:E1000,"80-84",Trades!J2:J1000,"WIN")/COUNTIFS(Trades!E2:E1000,"80-84",Trades!J2:J1000,"<>OPEN"),"")',
            ],
            [
                "80-84 avg R",
                '=IFERROR(AVERAGEIFS(Trades!L2:L1000,Trades!E2:E1000,"80-84",Trades!J2:J1000,"<>OPEN"),"")',
            ],
            ["85+ count", '=COUNTIF(Trades!E2:E1000,"85+")'],
            [
                "85+ WR",
                '=IFERROR(COUNTIFS(Trades!E2:E1000,"85+",Trades!J2:J1000,"WIN")/COUNTIFS(Trades!E2:E1000,"85+",Trades!J2:J1000,"<>OPEN"),"")',
            ],
            [
                "85+ avg R",
                '=IFERROR(AVERAGEIFS(Trades!L2:L1000,Trades!E2:E1000,"85+",Trades!J2:J1000,"<>OPEN"),"")',
            ],
            ["By direction", ""],
            ["LONG count", '=COUNTIF(Trades!C2:C1000,"LONG")'],
            [
                "LONG WR",
                '=IFERROR(COUNTIFS(Trades!C2:C1000,"LONG",Trades!J2:J1000,"WIN")/COUNTIFS(Trades!C2:C1000,"LONG",Trades!J2:J1000,"<>OPEN"),"")',
            ],
            ["SHORT count", '=COUNTIF(Trades!C2:C1000,"SHORT")'],
            [
                "SHORT WR",
                '=IFERROR(COUNTIFS(Trades!C2:C1000,"SHORT",Trades!J2:J1000,"WIN")/COUNTIFS(Trades!C2:C1000,"SHORT",Trades!J2:J1000,"<>OPEN"),"")',
            ],
        ]
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(
            writer, sheet_name="Summary", index=False, header=False
        )

    # 4. Inject execution formulas back across all active rows
    wb = load_workbook(excel_path)
    ws_trades = wb["Trades"]

    for row_idx in range(2, ws_trades.max_row + 1):
        # Column L: r_multiple formula
        ws_trades[
            f"L{row_idx}"
        ] = f'=IF(AND(K{row_idx}<>"",I{row_idx}<>"",G{row_idx}<>""),IF(C{row_idx}="LONG",(K{row_idx}-I{row_idx})/(I{row_idx}-G{row_idx}),(I{row_idx}-K{row_idx})/(G{row_idx}-I{row_idx})),"")'

        # Column M: slippage_r formula
        ws_trades[
            f"M{row_idx}"
        ] = f'=IF(AND(I{row_idx}<>"",F{row_idx}<>"",G{row_idx}<>""),IF(C{row_idx}="LONG",(I{row_idx}-F{row_idx})/(F{row_idx}-G{row_idx}),(F{row_idx}-I{row_idx})/(G{row_idx}-F{row_idx})),"")'

    wb.save(excel_path)
    print(
        print(f"  [Sync] Successfully updated '{excel_path}'. New trades appended, manual fields preserved.")
    )


if __name__ == "__main__":
    sync_json_to_excel()